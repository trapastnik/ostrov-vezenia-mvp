#!/usr/bin/env python3
"""
Импорт справочника ТН ВЭД ЕАЭС в базу данных.

Использование:
    # Из CSV файла (разделитель — табуляция или ;)
    python scripts/import_tn_ved.py data/tn_ved.csv

    # Из Excel файла (требует openpyxl)
    python scripts/import_tn_ved.py data/tn_ved.xlsx

    # Загрузить CSV с GitHub и импортировать
    python scripts/import_tn_ved.py --download

Формат CSV (табуляция или ;):
    Код;Наименование;Ед.изм.;Примечание

Источники данных:
    - TWS.BY (https://www.tws.by) — Excel, бесплатный
    - GitHub: https://github.com/nickleus27/russian_hs_codes
    - GitHub: https://github.com/AuraSci/hs-codes-csv
"""
import asyncio
import csv
import os
import sys

# Чтобы импортировать app.*, добавляем backend/ в PYTHONPATH
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.models.base import Base
from app.models.tn_ved_code import TnVedCode


def determine_level(code: str) -> int:
    """Определить уровень кода ТН ВЭД по длине значащей части."""
    stripped = code.rstrip("0")
    length = len(stripped)
    if length <= 2:
        return 2   # Группа (XX)
    elif length <= 4:
        return 4   # Позиция (XXXX)
    elif length <= 6:
        return 6   # Подпозиция (XXXXXX)
    elif length <= 8:
        return 8   # Подсубпозиция (XXXXXXXX)
    else:
        return 10  # Полный код (XXXXXXXXXX)


def determine_parent(code: str) -> str | None:
    """Определить родительский код."""
    stripped = code.rstrip("0")
    if len(stripped) <= 2:
        return None
    # Родитель — на 2 уровня выше
    parent_len = len(stripped) - 2
    if parent_len < 2:
        parent_len = 2
    parent_prefix = stripped[:parent_len]
    # Дополняем нулями до 10 знаков
    return parent_prefix.ljust(10, "0")


def parse_csv(filepath: str) -> list[dict]:
    """Прочитать CSV файл с кодами ТН ВЭД."""
    records = []

    with open(filepath, "r", encoding="utf-8-sig") as f:
        # Определяем разделитель
        sample = f.read(2048)
        f.seek(0)

        if "\t" in sample:
            delimiter = "\t"
        elif ";" in sample:
            delimiter = ";"
        else:
            delimiter = ","

        reader = csv.reader(f, delimiter=delimiter)

        for row in reader:
            if not row or len(row) < 2:
                continue

            code = row[0].strip().replace(" ", "").replace(".", "")

            # Пропускаем заголовок
            if not code or not code[0].isdigit():
                continue

            # Нормализуем код до 10 знаков
            code = code.ljust(10, "0")[:10]

            name = row[1].strip() if len(row) > 1 else ""
            unit = row[2].strip() if len(row) > 2 else None
            note = row[3].strip() if len(row) > 3 else None

            if not name:
                continue

            records.append({
                "code": code,
                "name": name,
                "level": determine_level(code),
                "parent_code": determine_parent(code),
                "unit": unit if unit else None,
                "note": note if note else None,
            })

    return records


def parse_excel(filepath: str) -> list[dict]:
    """Прочитать Excel файл с кодами ТН ВЭД (требует openpyxl)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Для импорта Excel установите: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    records = []

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue

        code = str(row[0]).strip().replace(" ", "").replace(".", "")
        if not code or not code[0].isdigit():
            continue

        code = code.ljust(10, "0")[:10]
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        unit = str(row[2]).strip() if len(row) > 2 and row[2] else None
        note = str(row[3]).strip() if len(row) > 3 and row[3] else None

        if not name:
            continue

        records.append({
            "code": code,
            "name": name,
            "level": determine_level(code),
            "parent_code": determine_parent(code),
            "unit": unit,
            "note": note,
        })

    wb.close()
    return records


async def import_records(records: list[dict]) -> int:
    """Импортировать записи в БД."""
    engine = create_async_engine(settings.DATABASE_URL, echo=False)

    # Создать таблицу если не существует
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async_session_factory = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with async_session_factory() as session:
        # Очистить старые данные
        await session.execute(text("DELETE FROM tn_ved_codes"))

        # Batch insert
        batch_size = 500
        imported = 0
        seen_codes = set()

        for i in range(0, len(records), batch_size):
            batch = records[i : i + batch_size]
            for record in batch:
                # Дедупликация
                if record["code"] in seen_codes:
                    continue
                seen_codes.add(record["code"])

                obj = TnVedCode(**record)
                session.add(obj)
                imported += 1

            await session.flush()
            print(f"  ... {imported} записей")

        await session.commit()

    await engine.dispose()
    return imported


def generate_demo_data() -> list[dict]:
    """Генерация демо-набора кодов ТН ВЭД (популярные категории)."""
    demo = [
        # Продукты
        ("0201", "Мясо крупного рогатого скота, свежее или охлаждённое", "кг", None),
        ("0207", "Мясо и субпродукты домашней птицы", "кг", None),
        ("0401", "Молоко и сливки, несгущённые", "л", None),
        ("0901", "Кофе", "кг", None),
        ("0902", "Чай", "кг", None),

        # Одежда (ДТЭГ часто)
        ("6101", "Пальто, полупальто мужские или для мальчиков, трикотажные", "шт", None),
        ("6102", "Пальто, полупальто женские или для девочек, трикотажные", "шт", None),
        ("6104", "Костюмы, комплекты, жакеты, платья женские, трикотажные", "шт", None),
        ("6109", "Футболки, майки трикотажные", "шт", None),
        ("610910", "Футболки, майки из хлопчатобумажной пряжи", "шт", None),
        ("6110", "Джемперы, пуловеры, кардиганы трикотажные", "шт", None),
        ("6203", "Костюмы, комплекты, пиджаки мужские", "шт", None),
        ("6204", "Костюмы, комплекты, жакеты, платья женские", "шт", None),

        # Обувь
        ("6401", "Водонепроницаемая обувь с подошвой и верхом из резины или пластмассы", "пар", None),
        ("6402", "Прочая обувь с подошвой и верхом из резины или пластмассы", "пар", None),
        ("6403", "Обувь с подошвой из резины, пластмассы, кожи и верхом из натуральной кожи", "пар", None),
        ("640399", "Прочая обувь с верхом из натуральной кожи", "пар", None),
        ("6404", "Обувь с подошвой из резины, пластмассы, кожи и верхом из текстильных материалов", "пар", None),

        # Электроника
        ("8471", "Вычислительные машины и их блоки", "шт", None),
        ("847130", "Портативные вычислительные машины массой не более 10 кг (ноутбуки)", "шт", None),
        ("847141", "Прочие вычислительные машины, содержащие процессор и устройства ввода/вывода", "шт", None),
        ("8517", "Аппаратура связи, телефонные аппараты, включая смартфоны", "шт", None),
        ("851712", "Телефоны для сотовых сетей связи (смартфоны)", "шт", None),
        ("8518", "Микрофоны, громкоговорители, наушники", "шт", None),
        ("8519", "Аппаратура звукозаписывающая или звуковоспроизводящая", "шт", None),
        ("8523", "Диски, ленты, устройства хранения данных (USB-флешки, SSD)", "шт", None),
        ("8528", "Мониторы и проекторы", "шт", None),

        # Косметика
        ("3303", "Духи и туалетная вода", "л", None),
        ("330300", "Духи и туалетная вода", "л", None),
        ("3304", "Средства для макияжа, ухода за кожей", "шт", None),
        ("3305", "Средства для волос", "шт", None),

        # Игрушки
        ("9503", "Трёхколёсные велосипеды, самокаты, игрушки прочие, модели", "шт", None),
        ("9504", "Игровые приставки и оборудование для развлечений", "шт", None),

        # Мебель
        ("9401", "Стулья и кресла", "шт", None),
        ("9403", "Мебель прочая и её части", "шт", None),
        ("940360", "Мебель деревянная прочая", "шт", None),

        # Бытовая техника
        ("8509", "Электромеханические бытовые машины с электродвигателем (миксеры, блендеры)", "шт", None),
        ("8510", "Электробритвы, машинки для стрижки, эпиляторы", "шт", None),
        ("8516", "Водонагреватели, электроутюги, фены, тостеры", "шт", None),

        # Автозапчасти
        ("8708", "Части и принадлежности моторных транспортных средств", "шт", None),

        # Книги
        ("4901", "Печатные книги, брошюры, листовки и аналогичные печатные материалы", "шт", None),

        # Спорт
        ("9506", "Оборудование для физкультуры, гимнастики, спорта", "шт", None),

        # Аксессуары
        ("4202", "Чемоданы, саквояжи, портфели, сумки", "шт", None),
        ("7117", "Бижутерия", "шт", None),

        # Часы
        ("9101", "Наручные часы, карманные часы с корпусом из драгоценных металлов", "шт", None),
        ("9102", "Наручные часы, карманные часы (прочие)", "шт", None),

        # Запрещённые (для валидации)
        ("2203", "Пиво солодовое (ЗАПРЕЩЕНО — подакцизный товар)", "л", "ЗАПРЕЩЕНО"),
        ("2204", "Вина виноградные (ЗАПРЕЩЕНО — подакцизный товар)", "л", "ЗАПРЕЩЕНО"),
        ("2208", "Спиртовые настойки, водка, ром, джин (ЗАПРЕЩЕНО — подакцизный товар)", "л", "ЗАПРЕЩЕНО"),
        ("2402", "Сигары, сигариллы, сигареты (ЗАПРЕЩЕНО — подакцизный товар)", "шт", "ЗАПРЕЩЕНО"),
    ]

    records = []
    for code, name, unit, note in demo:
        full_code = code.ljust(10, "0")[:10]
        records.append({
            "code": full_code,
            "name": name,
            "level": determine_level(full_code),
            "parent_code": determine_parent(full_code),
            "unit": unit,
            "note": note,
        })
    return records


async def main():
    if len(sys.argv) < 2:
        print("Использование:")
        print("  python scripts/import_tn_ved.py <файл.csv|файл.xlsx>")
        print("  python scripts/import_tn_ved.py --demo   # Загрузить демо-набор (~50 кодов)")
        sys.exit(1)

    arg = sys.argv[1]

    if arg == "--demo":
        print("Генерация демо-набора кодов ТН ВЭД...")
        records = generate_demo_data()
    elif arg.endswith(".xlsx") or arg.endswith(".xls"):
        print(f"Чтение Excel: {arg}")
        records = parse_excel(arg)
    else:
        print(f"Чтение CSV: {arg}")
        records = parse_csv(arg)

    print(f"Найдено {len(records)} кодов ТН ВЭД")

    if not records:
        print("Нет данных для импорта!")
        sys.exit(1)

    imported = await import_records(records)
    print(f"Импортировано {imported} кодов ТН ВЭД в базу данных")


if __name__ == "__main__":
    asyncio.run(main())
